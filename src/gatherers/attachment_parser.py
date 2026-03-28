"""
Email attachment parser — Phase 2, Sub-pipeline A (attachment handling).

Dispatches to the appropriate parser based on MIME type:
  - application/pdf              → PyMuPDF (fitz)
  - application/vnd...docx       → python-docx
  - application/vnd...xlsx       → openpyxl
  - image/*                      → placeholder (Claude Vision can be added)

All errors are caught and stored in AttachmentContent.error — never raise.
"""

from __future__ import annotations

import io
import logging
from typing import Optional

from src.schemas import AttachmentContent

logger = logging.getLogger(__name__)


def parse_attachment(
    filename: str,
    content: bytes,
    content_type: str,
) -> AttachmentContent:
    """
    Parse an email attachment and return extracted text.

    Args:
        filename: Original filename (used for fallback type detection).
        content: Raw bytes of the attachment.
        content_type: MIME type string from the email.

    Returns:
        AttachmentContent with extracted_text populated (or error set).
    """
    ct = content_type.lower()
    fn_lower = filename.lower()

    # Route by MIME type with filename fallback
    if "pdf" in ct or fn_lower.endswith(".pdf"):
        return _parse_pdf(filename, content)
    elif "wordprocessing" in ct or "msword" in ct or fn_lower.endswith((".docx", ".doc")):
        return _parse_docx(filename, content)
    elif "spreadsheet" in ct or "excel" in ct or fn_lower.endswith((".xlsx", ".xls", ".csv")):
        return _parse_excel(filename, content)
    elif ct.startswith("image/"):
        return _parse_image(filename, content, content_type)
    elif "text/" in ct or fn_lower.endswith((".txt", ".md")):
        return _parse_text(filename, content)
    else:
        return AttachmentContent(
            filename=filename,
            content_type=content_type,
            extracted_text="",
            extraction_method="none",
            error=f"Unsupported content type: {content_type}",
        )


def _parse_pdf(filename: str, content: bytes) -> AttachmentContent:
    """Extract text from PDF using PyMuPDF (fitz)."""
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(stream=content, filetype="pdf")
        total_pages = len(doc)  # cache before close
        pages_text: list[str] = []
        for page_num in range(total_pages):
            page = doc.load_page(page_num)
            text = page.get_text("text")
            if text.strip():
                pages_text.append(text.strip())
        doc.close()

        full_text = "\n\n".join(pages_text)
        if not full_text.strip():
            # PDF may be image-based — return empty with note
            return AttachmentContent(
                filename=filename,
                content_type="application/pdf",
                extracted_text="[PDF contains no extractable text — may be image-based]",
                page_count=total_pages,
                extraction_method="pymupdf",
            )

        return AttachmentContent(
            filename=filename,
            content_type="application/pdf",
            extracted_text=full_text,
            page_count=total_pages,
            extraction_method="pymupdf",
        )

    except ImportError:
        return AttachmentContent(
            filename=filename,
            content_type="application/pdf",
            extracted_text="",
            extraction_method="none",
            error="PyMuPDF (fitz) not installed — cannot parse PDF",
        )
    except Exception as e:
        logger.warning(f"PDF parsing failed for {filename}: {e}")
        return AttachmentContent(
            filename=filename,
            content_type="application/pdf",
            extracted_text="",
            extraction_method="pymupdf",
            error=str(e),
        )


def _parse_docx(filename: str, content: bytes) -> AttachmentContent:
    """Extract text from DOCX using python-docx."""
    try:
        from docx import Document

        doc = Document(io.BytesIO(content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        full_text = "\n\n".join(paragraphs)

        return AttachmentContent(
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            extracted_text=full_text,
            extraction_method="python-docx",
        )

    except ImportError:
        return AttachmentContent(
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            extracted_text="",
            extraction_method="none",
            error="python-docx not installed",
        )
    except Exception as e:
        logger.warning(f"DOCX parsing failed for {filename}: {e}")
        return AttachmentContent(
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            extracted_text="",
            extraction_method="python-docx",
            error=str(e),
        )


def _parse_excel(filename: str, content: bytes) -> AttachmentContent:
    """Extract cell text from XLSX/XLS using openpyxl or pandas (CSV fallback)."""
    fn_lower = filename.lower()

    if fn_lower.endswith(".csv"):
        try:
            import pandas as pd
            df = pd.read_csv(io.BytesIO(content), dtype=str)
            text = df.to_string(index=False)
            return AttachmentContent(
                filename=filename,
                content_type="text/csv",
                extracted_text=text,
                extraction_method="pandas",
            )
        except Exception as e:
            return AttachmentContent(
                filename=filename,
                content_type="text/csv",
                extracted_text="",
                extraction_method="pandas",
                error=str(e),
            )

    try:
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheets_text: list[str] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[str] = []
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    rows.append(row_text)
            if rows:
                sheets_text.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
        wb.close()

        return AttachmentContent(
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            extracted_text="\n\n".join(sheets_text),
            extraction_method="openpyxl",
        )

    except ImportError:
        return AttachmentContent(
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            extracted_text="",
            extraction_method="none",
            error="openpyxl not installed",
        )
    except Exception as e:
        logger.warning(f"Excel parsing failed for {filename}: {e}")
        return AttachmentContent(
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            extracted_text="",
            extraction_method="openpyxl",
            error=str(e),
        )


def _parse_image(filename: str, content: bytes, content_type: str) -> AttachmentContent:
    """
    Placeholder for image attachments.
    Future enhancement: pass to Claude Vision API for OCR.
    """
    return AttachmentContent(
        filename=filename,
        content_type=content_type,
        extracted_text="[Image attachment — Claude Vision OCR not yet configured]",
        extraction_method="none",
    )


def _parse_text(filename: str, content: bytes) -> AttachmentContent:
    """Parse plain text or markdown attachments."""
    try:
        text = content.decode("utf-8", errors="replace")
        return AttachmentContent(
            filename=filename,
            content_type="text/plain",
            extracted_text=text,
            extraction_method="decode",
        )
    except Exception as e:
        return AttachmentContent(
            filename=filename,
            content_type="text/plain",
            extracted_text="",
            extraction_method="decode",
            error=str(e),
        )
